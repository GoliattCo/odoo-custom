import base64
import io
import csv

from odoo import fields, models, _
from odoo.exceptions import UserError


class AccountMoveExportWizard(models.TransientModel):
    _name = 'account.move.export.wizard'
    _description = 'Exportar Asiento Contable a CSV/Excel'

    move_id = fields.Many2one(
        'account.move',
        string='Asiento Contable',
        required=True,
    )
    export_format = fields.Selection(
        [
            ('csv', 'CSV'),
            ('xlsx', 'Excel (XLSX)'),
        ],
        string='Formato',
        required=True,
        default='xlsx',
    )
    file_data = fields.Binary(string='Archivo', readonly=True)
    file_name = fields.Char(string='Nombre de Archivo', readonly=True)

    def _get_export_headers(self):
        return [
            'Código de Cuenta', 'Nombre de Cuenta', 'Etiqueta', 'Tercero',
            'Código Centro de Costo', 'Centro de Costo', 'Código de Ítem', 'Nombre de Ítem',
            'Auxiliar Abierto', 'Código Unidad de Negocio', 'Unidad de Negocio',
            'Débito', 'Crédito',
        ]

    def _get_export_rows(self):
        rows = []
        for line in self.move_id.line_ids.sorted(key=lambda l: l.sequence):
            if line.display_type in ('line_section', 'line_note'):
                continue
            rows.append([
                line.account_id.code or '',
                line.account_id.name or '',
                line.name or '',
                line.partner_id.name or '',
                line.cost_center_id.code or '',
                line.cost_center_id.name or '',
                line.item_code_id.code or '',
                line.item_code_id.name or '',
                line.auxiliar_abierto or '',
                line.business_unit_id.code or '',
                line.business_unit_id.name or '',
                line.debit,
                line.credit,
            ])
        return rows

    def action_export(self):
        self.ensure_one()
        headers = self._get_export_headers()
        rows = self._get_export_rows()
        entry_name = (self.move_id.name or 'Draft').replace('/', '_')

        if self.export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(rows)
            file_content = output.getvalue().encode('utf-8-sig')
            self.file_data = base64.b64encode(file_content)
            self.file_name = f'Asiento_{entry_name}.csv'
        else:
            try:
                import openpyxl
            except ImportError:
                raise UserError(_('La librería openpyxl es necesaria para exportar archivos Excel. Instálela con: pip install openpyxl'))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Asiento Contable'

            # Header row with bold formatting
            from openpyxl.styles import Font
            bold_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = bold_font

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

            output = io.BytesIO()
            wb.save(output)
            self.file_data = base64.b64encode(output.getvalue())
            self.file_name = f'Asiento_{entry_name}.xlsx'

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_move_id': self.move_id.id},
        }
